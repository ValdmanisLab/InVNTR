#!/usr/bin/env python3
import os
import argparse
import pandas as pd
from Bio.Seq import Seq
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
import pysam
import re
# For ANOVA and post-hoc
from scipy.stats import f_oneway
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import numpy as np

# -----------------------------
# Argument Parsing
# -----------------------------
def parse_args():
    parser = argparse.ArgumentParser(description='Extract VNTR')
    parser.add_argument('-f', '--folder', type=str, required=True, help='Path to folder')
    parser.add_argument('-l', '--length', type=int, nargs='?', const=1, default=0, help='Length of consensus motif. Takes priority over delimiter.')
    parser.add_argument('-d', '--delimiter', type=str, nargs='?', const=1, default='0', help='Delimiter marking start of motifs. Ignored if length is used.')
    parser.add_argument('-s', '--start', type=str, help='Beginning of VNTR')
    parser.add_argument('-e', '--end', type=str, help='End of VNTR')
    parser.add_argument('-ne', '--no_end', type=int, nargs='?', const=1, default=10000, help='If there is no end, cut after N bp.')
    parser.add_argument('-mal', '--max_allele_length', type=int, nargs='?', const=1, default=100000, help='Maximum allele length per file.')
    parser.add_argument('-t', '--type', type=str, nargs='?', const=1, default=0, help='Filetype (default .fa or .fasta)')
    parser.add_argument('-n', '--name', type=str, nargs='?', const=1, default='VNTR', help='Output prefix name')
    parser.add_argument(
        '-c', '--coordinate',
        type=str,
        help="Reference coordinates in genomic interval format, e.g. 'chr1:30908354-30908706'"
    )
    parser.add_argument('-r', '--reference', type=str, default='GRCh38', help='Reference genome (default GRCh38)')
    parser.add_argument('-p', '--print', action='store_true',
                    help='If set, export each sheet of the Excel output to CSV or TXT files.')
    parser.add_argument('--flank', type=int, default=50, help='Number of bp for flanking sequences (default 50)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose output')
    return parser.parse_args()

# -----------------------------
# Helpers
# -----------------------------
args = parse_args()
flank = args.flank  # assign to a local variable

log_messages = []

def timestamp():
    return datetime.now().strftime("%H:%M:%S")

def log(msg, step=None, record=True):
    if step:
        output = f"[{timestamp()}] [Step {step}] {msg}"
    else:
        output = f"[{timestamp()}] {msg}"
    print(output)
    if record:
        log_messages.append(output)

def read_fasta_file(filepath):
    with open(filepath) as f:
        return f.read().replace("\n", "")

def read_chromosome(fasta_file, chrom_name):
    """Return sequence for a single chromosome"""
    seq = []
    found = False
    with open(fasta_file) as f:
        for line in f:
            if line.startswith(">"):
                header = line.strip()[1:].split()[0]
                found = chrom_name in header
                continue
            if found:
                seq.append(line.strip())
            elif found and line.startswith(">"):
                break
    if not seq:
        raise RuntimeError(f"Chromosome {chrom_name} not found in {fasta_file}")
    return "".join(seq)

def check_variants_in_region(vcf_path, chrom, ref_coord_start, ref_coord_end):
    """
    Return a list of variants in a given region.
    Each variant is a tuple: (position, ref, alts)
    """
    variants = []
    try:
        vcf = pysam.VariantFile(vcf_path)
    except Exception as e:
        log(f"Could not open VCF: {vcf_path} ({e})")
        return variants

    try:
        for rec in vcf.fetch(chrom, ref_coord_start - 1, ref_coord_end):  # pysam uses 0-based coords
            variants.append((rec.pos, rec.ref, ",".join(rec.alts)))
    except ValueError as e:
        log(f"VCF fetch error for {chrom}:{ref_coord_start}-{ref_coord_end}: {e}")
    finally:
        vcf.close()

    return variants

def create_nucleotide_df(seq, genome_start_coord):
    """
    Create a dataframe with each nucleotide and its genomic coordinate.
    genome_start_coord: 0-based start position of seq in genome
    Returns: pandas DataFrame with columns ['Pos', 'Base']
    """
    df = pd.DataFrame({
        'Pos': range(genome_start_coord + 1, genome_start_coord + 1 + len(seq)),  # 1-based coordinates
        'Base': list(seq)
    })
    return df


def annotate_sequence_with_variants(nuc_df, var_df):
    """
    Annotate a nucleotide sequence (from nuc_df) with variants (from var_df).
    nuc_df: DataFrame with columns ['Pos', 'Base']
    var_df: DataFrame with columns ['Pos', 'REF', 'ALT']
    Returns annotated sequence as a string, e.g. C[>T]A[>G]...
    """
    seq = list(nuc_df['Base'])
    annotations = [""] * len(seq)

    for _, row in var_df.iterrows():
        pos = row['Pos']
        ref = row['REF']
        alt = row['ALT']
        idx = pos - nuc_df['Pos'].iloc[0]  # 0-based index
        if 0 <= idx < len(seq):
            if seq[idx].upper() == ref.upper():
                annotations[idx] = f"[>{alt}]"
            else:
                # mismatch: still annotate
                annotations[idx] = f"[>{alt}]"

    annotated_seq = "".join(base + ann for base, ann in zip(seq, annotations))
    return annotated_seq


def mask_variant_positions(annotated_seq):
    """
    Convert a variant-annotated sequence into a regex string where
    all variant positions are replaced with '.' (any nucleotide) and
    all other bases remain.
    """
    import re
    # Replace any base followed by [>...] with '.'
    regex_seq = re.sub(r'([ACGTacgt])\[>.*?\]', '.', annotated_seq)
    return regex_seq


def reverse_complement(seq):
    """Return the reverse complement of a DNA sequence."""
    complement = str.maketrans('ACGTacgt', 'TGCAtgca')
    return seq.translate(complement)[::-1]

def safe_search(pattern, seq):
    """Return the start index of the regex pattern in seq, or -1 if not found."""
    if not pattern:
        return -1
    m = re.search(pattern, seq)
    return m.start() if m else -1
# -----------------------------
# Reference Loading
# -----------------------------
def load_reference_info(folder, ref_name):
    ref_aliases = {
        'GRCh38': ['grch38', 'hg38', '38'],
        'CHM13': ['chm13', 't2t']
    }
    
    ref_found = None
    for ref_key, aliases in ref_aliases.items():
        if ref_name.lower() in aliases or ref_name.lower() == ref_key.lower():
            ref_found = ref_key
            break
    
    if not ref_found:
        raise ValueError(f"Unknown reference: {ref_name}")
    
    fasta_file = os.path.join(folder, f"{ref_found}.fa")
    fai_file = os.path.join(folder, f"{ref_found}.fa.fai")
    
    if not os.path.exists(fasta_file) or not os.path.exists(fai_file):
        raise FileNotFoundError(f"Reference {ref_found} FASTA or FAI not found in folder {folder}")
    
    chroms = {}
    with open(fai_file) as f:
        for line in f:
            fields = line.strip().split("\t")
            chrom_name = fields[0].split("#")[-1]
            chrom_len = int(fields[1])
            chroms[chrom_name] = chrom_len
    
    return fasta_file, chroms, ref_found

# -----------------------------
# VNTR Extraction Function
# -----------------------------
def extract_vntr(sequence, five_prime_seq, three_prime_seq, filename, args,
                 five_prime_regex=None, five_prime_regex_rc=None,
                 three_prime_regex=None, three_prime_regex_rc=None,
                 flank=None):
    """
    Extract the VNTR sequence between five_prime_seq and three_prime_seq in 'sequence'.
    Handles both forward and reverse-complement matches, with regex optional.
    Returns: (VNTR sequence or None, list of errors)
    """
    errors = []

    # Precompute reverse complements of flanking sequences
    five_prime_rc = str(Seq(five_prime_seq).reverse_complement())
    three_prime_rc   = str(Seq(three_prime_seq).reverse_complement())

    # Special case: no end sequence
    if not three_prime_seq or three_prime_seq.lower() == 'n':
        five_prime_index = safe_search(five_prime_regex, sequence) if five_prime_regex else sequence.find(five_prime_seq)
        if five_prime_index >= 0:
            c = sequence[five_prime_index + flank : five_prime_index + flank + args.no_end]  # skip 5' flank
            return c, errors
        # Try reverse complement
        five_prime_index_rc = safe_search(five_prime_regex_rc, sequence) if five_prime_regex_rc else sequence.find(five_prime_rc)
        if five_prime_index_rc >= 0:
            c = sequence[max(0, five_prime_index_rc - args.no_end) : five_prime_index_rc]  # already reverse handled, no need to skip 5' flank
            return str(Seq(c).reverse_complement()), errors

        errors.append(f"{filename}: No sequence found")
        return None, errors

    # Standard extraction: forward
    five_prime_index = safe_search(five_prime_regex, sequence) if five_prime_regex else sequence.find(five_prime_seq)
    three_prime_index = safe_search(three_prime_regex, sequence) if three_prime_regex else sequence.find(three_prime_seq)

    if five_prime_index >= 0 and three_prime_index >= 0:
        if three_prime_index < five_prime_index:
            errors.append(f"{filename}: unexpected order for indexes, forward 5' found after forward 3'")
            return None, errors
        c = sequence[five_prime_index + flank : three_prime_index]  # adjust for flank in 5' index
        return c, errors

    # Reverse-complement extraction
    five_prime_index_rc = safe_search(five_prime_regex_rc, sequence) if five_prime_regex_rc else sequence.find(five_prime_rc)
    three_prime_index_rc = safe_search(three_prime_regex_rc, sequence) if three_prime_regex_rc else sequence.find(three_prime_rc)

    if five_prime_index_rc >= 0 and three_prime_index_rc >= 0:
        if five_prime_index_rc < three_prime_index_rc:
            errors.append(f"{filename}: unexpected order for indexes, reverse 3' found after reverse 5'")
            return None, errors
        c = sequence[three_prime_index_rc + flank:five_prime_index_rc:]  # adjust for flank in 5' rc index
        return str(Seq(c).reverse_complement()), errors


    # Handle partial matches
    if five_prime_index >= 0 and three_prime_index < 0:
        errors.append(f"{filename}: only 5' sequence found, no 3'")
    elif five_prime_index < 0 and three_prime_index >= 0:
        errors.append(f"{filename}: only 3' sequence found, no 5'")
    elif five_prime_index_rc >= 0 and three_prime_index_rc < 0:
        errors.append(f"{filename}: only reverse complement 5' sequence found, no reverse complement 3'")
    elif five_prime_index_rc < 0 and three_prime_index_rc >= 0:
        errors.append(f"{filename}: only reverse complement 3' sequence found, no reverse complement 5'")
    else:
        errors.append(f"{filename}: no matches to 3' or 5' sequences found")

    return None, errors


    
# -----------------------------
# Motif Split
# -----------------------------
def split_into_motifs(seq, motif_length, delimiter):
    if motif_length and motif_length > 0:
        return [seq[i:i + motif_length] for i in range(0, len(seq), motif_length)]
    elif delimiter != '0':
        return [delimiter + e for e in seq.rsplit(delimiter) if e]
    else:
        return []

# -----------------------------
# Output Helpers
# -----------------------------
def record_outputs(alleles, allele_length, filename, vntr_seq, maxlen):
    alleles.append(f"{filename}\n{vntr_seq[:maxlen]}\n")
    allele_length.append(f"{filename},{len(vntr_seq[:maxlen])}")


# -----------------------------
# Worker Function for Parallelism (Updated for regex search with variants)
# -----------------------------
def process_file(args, filename, five_prime_seq, three_prime_seq,
                 five_prime_regex=None, five_prime_regex_rc=None,
                 three_prime_regex=None, three_prime_regex_rc=None):
    filepath = os.path.join(args.folder, filename)
    seq = read_fasta_file(filepath)
    
    # Call extract_vntr with optional masked regex sequences
    vntr_seq, errs = extract_vntr(
        sequence=seq,
        five_prime_seq=five_prime_seq,
        three_prime_seq=three_prime_seq,
        filename=filename,
        args=args,
        five_prime_regex=five_prime_regex,
        five_prime_regex_rc=five_prime_regex_rc,
        three_prime_regex=three_prime_regex,
        three_prime_regex_rc=three_prime_regex_rc,
        flank=flank
    )
    
    motifs = split_into_motifs(vntr_seq, args.length, args.delimiter) if vntr_seq else []
    return filename, vntr_seq, motifs, errs

# -----------------------------
# Write Outputs
# -----------------------------
def write_outputs(df, alleles, allele_length, errors, outputname, assemblies_meta, print_sheets=False):
    df_clean = df.dropna(how='all')    
    freq = df_clean.stack().value_counts()
    
    excel_name = f"{outputname}.xlsx"
    counter = 1
    while os.path.exists(excel_name):
        excel_name = f"{outputname}_{counter}.xlsx"
        counter += 1
    
    wb = Workbook()
    
    # SeattlePlot sheet (renamed from VNTR)
    ws_vntr = wb.active
    ws_vntr.title = "SeattlePlot"
    
    # First, write just the header row
    header_written = False
    motif_data_rows = []
    
    for r in dataframe_to_rows(df_clean, index=True, header=True):
        # r is a tuple/list of cell values for the row; skip if all None or empty string
        if all((c is None or (isinstance(c, str) and c.strip() == "")) for c in r):
            continue
        
        if not header_written:
            # Write the header row first
            ws_vntr.append(r)
            header_written = True
        else:
            # Store motif data rows for later
            motif_data_rows.append(r)
    
    # Build a quick map from allele_length list: "filename,length"
    allele_len_map = {}
    for line in allele_length:
        parts = line.split(",")
        if len(parts) >= 2:
            try:
                allele_len_map[parts[0]] = int(parts[1])
            except ValueError:
                allele_len_map[parts[0]] = parts[1]
    
    # Insert metadata rows right after the header (starting at row 2)
    labels = ["Allele Length", "Population code", "Superpopulation code", "File Origin"]
    
    # Write labels in column A and fill values for each filename column
    for i, lab in enumerate(labels):
        metadata_row = [lab]  # Start with the label in column A
        
        # Fill values for each filename column (filenames are in the header row starting from col 2)
        for col_idx in range(2, ws_vntr.max_column + 1):
            header_val = ws_vntr.cell(row=1, column=col_idx).value
            if not header_val:
                metadata_row.append("")
                continue
            
            filename = str(header_val)
            
            if i == 0:  # Allele Length
                metadata_row.append(allele_len_map.get(filename, "NA"))
            elif i == 1:  # Population code
                meta = assemblies_meta.get(filename, {})
                metadata_row.append(meta.get("Population code", "NA"))
            elif i == 2:  # Superpopulation code
                meta = assemblies_meta.get(filename, {})
                metadata_row.append(meta.get("Superpopulation code", "NA"))
            elif i == 3:  # File Origin
                meta = assemblies_meta.get(filename, {})
                metadata_row.append(meta.get("File Origin", "NA"))
        
        ws_vntr.append(metadata_row)
    
    # Now append all the motif data rows
    for row in motif_data_rows:
        ws_vntr.append(row)

    ws_vntr.cell(row=1, column=1, value="Filename")
    
    # Allele_Length sheet
    ws_allele_len = wb.create_sheet(title="Allele_Length")
    # Header row
    ws_allele_len.append(["Filename", "Allele_Length", "Population code", "Superpopulation code", "File Origin"])
    
    for line in allele_length:
        parts = line.split(",")
        if len(parts) < 2:
            continue
        filename, length = parts[0], parts[1]
        try:
            length_val = int(length)
        except ValueError:
            length_val = length
        
        meta = assemblies_meta.get(filename, {})
        ws_allele_len.append([
            filename,
            length_val,
            meta.get("Population code", "NA"),
            meta.get("Superpopulation code", "NA"),
            meta.get("File Origin", "NA")
        ])
    
    # Frequency sheet
    ws_freq = wb.create_sheet(title="Frequency")
    ws_freq.append(["Motif", "Frequency", "Length"])
    for val, count in freq.items():
        motif = val
        ws_freq.append([motif, int(count), len(str(motif))])
    
    # Alleles sheet (FASTA-like)
    ws_alleles = wb.create_sheet(title="Alleles")
    row = 1
    for item in alleles:
        filename, seq = item.strip().split("\n")
        ws_alleles.cell(row=row, column=1, value=f">{filename}")
        row += 1
        ws_alleles.cell(row=row, column=1, value=seq)
        row += 2
    
    # Errors sheet
    ws_errors = wb.create_sheet(title="Errors")
    ws_errors.append(["Sample", "Error"])
    for err in errors:
        if ": " in err:
            sample, msg = err.split(": ", 1)
        else:
            sample, msg = "", err
        ws_errors.append([sample, msg])
    
    # Population sheet: means and ANOVA/Tukey
    ws_population = wb.create_sheet(title="Population")
    
    # Build a DataFrame from allele_length and assemblies_meta
    rows = []
    for line in allele_length:
        parts = line.split(",")
        if len(parts) < 2:
            continue
        filename = parts[0]
        try:
            length_val = int(parts[1])
        except ValueError:
            # skip non-integer lengths
            continue
        
        meta = assemblies_meta.get(filename, {})
        pop_code = meta.get("Population code", "") or "NA"
        superpop_code = meta.get("Superpopulation code", "") or "NA"
        
        rows.append({
            "Filename": filename,
            "Allele_Length": length_val,
            "Population code": pop_code,
            "Superpopulation code": superpop_code
        })
    
    pop_df = pd.DataFrame(rows)
    
    # Means per Population code
    ws_population.append(["Population code", "N", "Mean_Allele_Length", "Std_Allele_Length"])
    if not pop_df.empty:
        pop_grouped = pop_df.groupby("Population code")["Allele_Length"].agg(['count', 'mean', 'std']).reset_index()
        for _, r in pop_grouped.iterrows():
            ws_population.append([
                r['Population code'],
                int(r['count']),
                float(r['mean']),
                float(r['std']) if not np.isnan(r['std']) else "NA"
            ])
    else:
        ws_population.append(["No data", "", "", ""])
    
    ws_population.append([])
    
    # Means per Superpopulation code
    ws_population.append(["Superpopulation code", "N", "Mean_Allele_Length", "Std_Allele_Length"])
    if not pop_df.empty:
        super_grouped = pop_df.groupby("Superpopulation code")["Allele_Length"].agg(['count', 'mean', 'std']).reset_index()
        for _, r in super_grouped.iterrows():
            ws_population.append([
                r['Superpopulation code'],
                int(r['count']),
                float(r['mean']),
                float(r['std']) if not np.isnan(r['std']) else "NA"
            ])
    else:
        ws_population.append(["No data", "", "", ""])
    
    ws_population.append([])
    
    # ANOVA & Tukey for Population code
    ws_population.append(["ANOVA - Population code"])
    pop_significant = False
    if pop_df.empty:
        ws_population.append(["No allele length data available for ANOVA"])
    else:
        # Prepare groups
        pop_groups = {g: grp["Allele_Length"].values for g, grp in pop_df.groupby("Population code")}
        # Filter out groups with zero length (shouldn't happen)
        pop_groups = {k: v for k, v in pop_groups.items() if len(v) > 0}
        
        if len(pop_groups) < 2:
            ws_population.append(["Not enough groups for ANOVA (need >=2)"])
        else:
            try:
                anova_res = f_oneway(*[v for v in pop_groups.values()])
                ws_population.append(["F-statistic", anova_res.statistic])
                ws_population.append(["p-value", anova_res.pvalue])
                
                # Check for significance
                if anova_res.pvalue < 0.05:
                    pop_significant = True
                    
            except Exception as e:
                ws_population.append(["ANOVA error", str(e)])
    
    # Tukey HSD
    try:
        data = pop_df["Allele_Length"].values
        groups = pop_df["Population code"].values
        tuk = pairwise_tukeyhsd(endog=data, groups=groups, alpha=0.05)
        
        ws_population.append([])
        ws_population.append(["Tukey HSD - Population code"])
        
        # Tukey result summary as lines
        tuk_summary = tuk.summary()
        # tuk.summary() returns an object with .data (list of rows)
        for row in tuk_summary.data:
            ws_population.append(list(row))
            
    except Exception as e:
        ws_population.append(["Tukey HSD error", str(e)])
    
    ws_population.append([])
    
    # ANOVA & Tukey for Superpopulation code
    ws_population.append(["ANOVA - Superpopulation code"])
    superpop_significant = False
    if pop_df.empty:
        ws_population.append(["No allele length data available for ANOVA"])
    else:
        super_groups = {g: grp["Allele_Length"].values for g, grp in pop_df.groupby("Superpopulation code")}
        super_groups = {k: v for k, v in super_groups.items() if len(v) > 0}
        
        if len(super_groups) < 2:
            ws_population.append(["Not enough groups for ANOVA (need >=2)"])
        else:
            try:
                anova_res2 = f_oneway(*[v for v in super_groups.values()])
                ws_population.append(["F-statistic", anova_res2.statistic])
                ws_population.append(["p-value", anova_res2.pvalue])
                
                # Check for significance
                if anova_res2.pvalue < 0.05:
                    superpop_significant = True
                    
            except Exception as e:
                ws_population.append(["ANOVA error", str(e)])
    
    # Tukey HSD
    try:
        data2 = pop_df["Allele_Length"].values
        groups2 = pop_df["Superpopulation code"].values
        tuk2 = pairwise_tukeyhsd(endog=data2, groups=groups2, alpha=0.05)
        
        ws_population.append([])
        ws_population.append(["Tukey HSD - Superpopulation code"])
        
        tuk2_summary = tuk2.summary()
        for row in tuk2_summary.data:
            ws_population.append(list(row))
            
    except Exception as e:
        ws_population.append(["Tukey HSD error", str(e)])
    
    # Log significance results
    if pop_significant or superpop_significant:
        if pop_significant and superpop_significant:
            log("Significant differences found between populations and between superpopulations")
        elif pop_significant:
            log("Significant differences found between populations")
        elif superpop_significant:
            log("Significant differences found between superpopulations")
    else:
        log("No significant differences found between populations or superpopulations")
    
    # Log sheet
    ws_log = wb.create_sheet(title="Log")
    ws_log.cell(row=1, column=1, value=datetime.today().strftime("%Y-%m-%d"))
    ws_log.cell(row=1, column=2, value=f"{sys.executable} " + " ".join(sys.argv))
    
    for i, line in enumerate(log_messages, start=2):
        if line.startswith("["):
            ts_end = line.find("]")
            timestamp_part = line[1:ts_end]
            message_part = line[ts_end+2:]
        else:
            timestamp_part = ""
            message_part = line
        ws_log.cell(row=i, column=1, value=timestamp_part)
        ws_log.cell(row=i, column=2, value=message_part)
    
    # Compute total elapsed between first and last timestamp in the log and average per genome
    # Find first and last timestamp entries in log_messages
    times = []
    for line in log_messages:
        if line.startswith("["):
            ts_end = line.find("]")
            tstr = line[1:ts_end]
            try:
                t = datetime.strptime(tstr, "%H:%M:%S")
                times.append(t)
            except Exception:
                # ignore unparsable
                pass
    
    elapsed_seconds = 0.0
    avg_per_genome = 0.0
    num_genomes = len(allele_length) if allele_length else 0
    
    if len(times) >= 2 and num_genomes > 0:
        first = times[0]
        last = times[-1]
        # Compute difference in seconds, accounting for possible day wrap
        delta = (last - first).total_seconds()
        if delta < 0:  # wrapped past midnight, add 24h
            delta += 24 * 3600
        elapsed_seconds = delta
        avg_per_genome = elapsed_seconds / num_genomes
    elif len(times) >= 2:
        first = times[0]
        last = times[-1]
        delta = (last - first).total_seconds()
        if delta < 0:
            delta += 24 * 3600
        elapsed_seconds = delta
        avg_per_genome = float('nan')  # no genomes to divide by
    
    ws_log.append([])
    ws_log.append(["Total elapsed time (seconds between first and last log timestamp):", f"{elapsed_seconds:.4f}"])
    ws_log.append(["Number of genomes processed (allele_length entries):", f"{num_genomes}"])
    ws_log.append(["Average time per genome (seconds):", f"{avg_per_genome:.4f}" if not np.isnan(avg_per_genome) else "NA"])
    ws_log.append(["Completed at:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    if print_sheets:  # pass as a parameter to write_outputs()
        base = outputname

        for sheet_name, ext in [
            ("SeattlePlot", "csv"),
            ("Allele_Length", "csv"),
            ("Frequency", "csv"),
            ("Population", "csv")
        ]:
            ws = wb[sheet_name]
            data = [[cell.value for cell in row] for row in ws.rows]
            pd.DataFrame(data).to_csv(f"{base}_{sheet_name}.{ext}", index=False, header=False)

        for sheet_name, ext in [
            ("Alleles", "txt"),
            ("Errors", "txt"),
            ("Log", "txt")
        ]:
            ws = wb[sheet_name]
            with open(f"{base}_{sheet_name}.{ext}", "w") as f:
                for row in ws.iter_rows(values_only=True):
                    line = "\t".join([str(c) for c in row if c is not None])
                    if line.strip():
                        f.write(line + "\n")

        log("All Excel sheets exported to individual CSV/TXT files due to -p / --print")

    # After writing the Excel file (or right before finishing the function)
    num_errors = len(errors)
    if num_errors > 0:
        log(f"Total errors encountered: {num_errors}")
    else:
        log("No errors encountered.")

    wb.save(excel_name)
    log(f"Excel output written to {excel_name}")




# -----------------------------
# Main Workflow
# -----------------------------
def main():
    args = parse_args()
    step_counter = 1
    
    log("Parsing arguments...", step_counter)
    step_counter += 1
    
    if args.coordinate:
        fasta_file, chroms, ref_name = load_reference_info(args.folder, args.reference)
        log(f"Using reference genome: {ref_name} ({fasta_file})", step_counter)
        step_counter += 1
    
    # Load assemblies.csv (comma-separated)
    assemblies_path = os.path.join(args.folder, "assemblies.csv")
    if not os.path.exists(assemblies_path):
        raise FileNotFoundError(f"assemblies.csv not found in {args.folder}")
    
    assemblies_df = pd.read_csv(assemblies_path, sep=",", dtype=str).fillna("")
    
    if "Filename" not in assemblies_df.columns:
        raise RuntimeError("assemblies.csv must contain a 'Filename' column")
    
    assemblies_meta = {row["Filename"]: row.to_dict() for _, row in assemblies_df.iterrows()}
    log(f"Loaded metadata for {len(assemblies_meta)} assemblies.", step_counter)
    step_counter += 1
    
    if args.coordinate:
        chr_name, region = args.coordinate.split(":")
        five_prime_coord, three_prime_coord = [int(x.replace(",", "")) for x in region.split("-")]

        chrom_seq = read_chromosome(fasta_file, chr_name)
        chrom_len = len(chrom_seq)
        flank = args.flank

        # Slice flanking sequences directly
        five_prime_seq = chrom_seq[max(0, five_prime_coord - 1 - flank): five_prime_coord - 1]
        three_prime_seq = chrom_seq[three_prime_coord: min(chrom_len, three_prime_coord + flank)]

        log(f"Chromosome {chr_name} loaded ({chrom_len} bp)")
        log(f"Extracted {chr_name}:{five_prime_coord}-{three_prime_coord} ±{flank}bp")
        log(f"5' flank ({len(five_prime_seq)}bp): {five_prime_seq}")
        log(f"3' flank ({len(three_prime_seq)}bp): {three_prime_seq}")

        # Create nucleotide DataFrames for variant annotation
        five_prime_df = create_nucleotide_df(five_prime_seq, max(0, five_prime_coord - 1 - flank))
        three_prime_df = create_nucleotide_df(three_prime_seq, three_prime_coord)

        # --- Variant lookup in min VCF ---
        vcf_path = os.path.join(args.folder, "ALL.sites.combined.min.vcf.gz")

        masked_five_prime_seq = None
        masked_five_prime_seq_rc = None
        masked_three_prime_seq = None
        masked_three_prime_seq_rc = None

        if os.path.exists(vcf_path):
            # 5' flank variants
            five_prime_flank_variants = check_variants_in_region(
                vcf_path,
                chr_name,
                max(0, five_prime_coord - 1 - flank) + 1,
                five_prime_coord - 1
            )
            # 3' flank variants
            three_prime_flank_variants = check_variants_in_region(
                vcf_path,
                chr_name,
                three_prime_coord + 1,
                min(chrom_len, three_prime_coord + flank)
            )

            # Convert to DataFrames
            five_prime_var_df = pd.DataFrame(five_prime_flank_variants, columns=["Pos", "REF", "ALT"])
            three_prime_var_df = pd.DataFrame(three_prime_flank_variants, columns=["Pos", "REF", "ALT"])

            # Mask sequences if variants exist
            if not five_prime_var_df.empty:
                masked_five_prime_seq = mask_variant_positions(
                    annotate_sequence_with_variants(five_prime_df, five_prime_var_df)
                )
                masked_five_prime_seq_rc = reverse_complement(masked_five_prime_seq)

            if not three_prime_var_df.empty:
                masked_three_prime_seq = mask_variant_positions(
                    annotate_sequence_with_variants(three_prime_df, three_prime_var_df)
                )
                masked_three_prime_seq_rc = reverse_complement(masked_three_prime_seq)

        else:
            log(f"VCF file not found: {vcf_path}. Variant masking skipped for flanks.")

        log("Reference VNTR extraction prepared.", step_counter)
        step_counter += 1

    else:
        # --- Manual start/end input mode ---
        if not args.start:
            raise ValueError("Either --coordinate or --start must be provided.")
        five_prime_seq = args.start
        three_prime_seq = args.end if args.end else "n"
        log(f"Using manual boundary sequences.", step_counter)
        log(f"5' sequence: {five_prime_seq}")
        log(f"3' sequence: {three_prime_seq}")
        step_counter += 1

    # --- Process genome files ---
    files = [f for f in os.listdir(args.folder) if f.endswith((".fa", ".fasta"))]
    df = pd.DataFrame(index=range(1, 10000))
    alleles, allele_length, errors = [], [], []
    
    log("Processing genome files (multithreaded)...", step_counter)
    total_files = len(files)
    completed = 0
    
    # --- Parallel processing ---
    with ProcessPoolExecutor(max_workers=8) as executor:
        futures = {}
        for f in files:
            futures[executor.submit(
                process_file,
                args,
                f,
                five_prime_seq,
                three_prime_seq,
                five_prime_regex=masked_five_prime_seq if 'masked_five_prime_seq' in locals() else None,
                five_prime_regex_rc=masked_five_prime_seq_rc if 'masked_five_prime_seq_rc' in locals() else None,
                three_prime_regex=masked_three_prime_seq if 'masked_three_prime_seq' in locals() else None,
                three_prime_regex_rc=masked_three_prime_seq_rc if 'masked_three_prime_seq_rc' in locals() else None
            )] = f

        for future in as_completed(futures):
            filename = futures[future]
            try:
                filename, vntr_seq, motifs, errs = future.result()
                errors.extend(errs)
                if vntr_seq:
                    all_motifs[filename] = pd.Series(motifs)
                    record_outputs(alleles, allele_length, filename, vntr_seq, args.max_allele_length)
            except Exception as e:
                errors.append(f"{filename}: {e}")
            
            completed += 1
            filled = int(50 * completed / total_files) if total_files else 50
            bar = "#" * filled + "-" * (50 - filled)
            print(f"\rProcessing genomes: [{bar}] {completed}/{total_files}", end="")
    
    print()
    df = pd.DataFrame(all_motifs)
    df.index += 1
    write_outputs(df, alleles, allele_length, errors, args.name, assemblies_meta, print_sheets=args.print)




all_motifs = {}
if __name__ == "__main__":
    main()
