pip install openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

# Create a sample DataFrame
data = {'Column1': [1, 2, 3, 4, 5],
        'Column2': ['A', 'B', 'C', 'D', 'E'],
        'Column3': ['Yellow', 'Green', 'Red', 'Yellow', 'Green']}
df = pd.DataFrame(data)

# Create a Workbook object
writer = pd.ExcelWriter('data.xlsx', engine='openpyxl')
writer.book = writer.book = writer.book = writer.book

# Convert DataFrame to Excel
df.to_excel(writer, index=False, sheet_name='Sheet1')

# Get the Workbook and Sheet objects
workbook = writer.book
sheet = writer.sheets['Sheet1']

# Iterate over each row and apply color based on matching values
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=df.shape[0] + 1, max_col=df.shape[1]):
    search_value = row[0].value
    color_value = row[2].value
    
    fill = PatternFill(start_color=color_value, end_color=color_value, fill_type="solid")
    
    for cell in row:
        if cell.value == search_value:
            cell.fill = fill

# Save the changes and close the Workbook
writer.save()
writer.close()